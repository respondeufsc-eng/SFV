# app.py - VERSÃO COMPLETA CORRIGIDA
from flask import Flask, send_file, request, render_template, flash, session, jsonify, redirect
from flask_session import Session  # ADICIONADO: Importação necessária
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side, NamedStyle, Protection
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
import pandas as pd
import tempfile
import io
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm, inch
from reportlab.graphics.shapes import Drawing
from reportlab.graphics.charts.barcharts import VerticalBarChart
from reportlab.graphics.charts.piecharts import Pie
from reportlab.graphics import renderPDF
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from datetime import datetime
import numpy as np
import math
import json
import traceback
import warnings
import os
import re
warnings.filterwarnings('ignore')

_UNSUPPORTED_CHARS_RE = re.compile(
    u'['
    u'\U00010000-\U0010FFFF'   # Non-BMP: 🔧 and other emoji
    u'\U00002600-\U000027BF'   # Misc Symbols + Dingbats: ✅ ♻ ⚠
    u'\U0000FE00-\U0000FE0F'   # Variation selectors (e.g. the invisible ️ after ♻)
    u']',
    re.UNICODE
)

def _pdf_text(text):
    """Strip characters Arial cannot render (shows as □ in PDF)."""
    return _UNSUPPORTED_CHARS_RE.sub('', str(text)).strip()

_pdf_fonts_registered = False

def _register_pdf_fonts():
    global _pdf_fonts_registered
    if _pdf_fonts_registered:
        return
    fonts_dir = os.path.join(os.environ.get('WINDIR', 'C:\\Windows'), 'Fonts')
    try:
        pdfmetrics.registerFont(TTFont('Arial', os.path.join(fonts_dir, 'arial.ttf')))
        pdfmetrics.registerFont(TTFont('Arial-Bold', os.path.join(fonts_dir, 'arialbd.ttf')))
        pdfmetrics.registerFont(TTFont('Arial-Italic', os.path.join(fonts_dir, 'ariali.ttf')))
        pdfmetrics.registerFont(TTFont('Arial-BoldItalic', os.path.join(fonts_dir, 'arialbi.ttf')))
        from reportlab.pdfbase.pdfmetrics import registerFontFamily
        registerFontFamily('Arial', normal='Arial', bold='Arial-Bold', italic='Arial-Italic', boldItalic='Arial-BoldItalic')
        _pdf_fonts_registered = True
    except Exception as e:
        print(f'Aviso: Nao foi possivel registrar fontes Arial: {e}')

app = Flask(__name__, static_folder='static', template_folder='templates')
app.secret_key = 'segredo_ultra_seguro_2025'

# CONFIGURAÇÃO DE SESSÃO CORRIGIDA
app.config['SESSION_TYPE'] = 'filesystem'
app.config['MAX_CONTENT_LENGTH'] = 50 * 1024 * 1024  # 50MB max file size
Session(app)  # ADICIONADO: Inicializa a sessão no sistema de arquivos

# Função auxiliar para converter strings numéricas com ponto ou vírgula
def converter_numero(valor):
    """Converte string numérica com ponto ou vírgula para float"""
    if pd.isna(valor):
        return 0.0
    
    try:
        if isinstance(valor, str):
            # Remover espaços e substituir vírgula por ponto
            valor_limpo = valor.strip().replace(',', '.')
            
            # Verificar se é erro do Excel
            if any(error in valor_limpo.upper() for error in ['#DIV/0!', '#N/A', '#VALUE!', '#REF!', '#NAME?', 'N/A']):
                return 0.0
            
            # Remover qualquer caractere não numérico (exceto ponto e sinal negativo)
            valor_limpo = ''.join(c for c in valor_limpo if c.isdigit() or c in '.-')
            
            if valor_limpo and valor_limpo != '-':
                return float(valor_limpo)
            else:
                return 0.0
        elif isinstance(valor, (int, float)):
            return float(valor)
        else:
            return 0.0
    except:
        return 0.0

# Função SEGURA para converter DataFrame para dicionário serializável
def dataframe_para_dict_serializavel_seguro(df):
    """Converte DataFrame para dicionário serializável de forma SEGURA"""
    dados_serializaveis = {}
    
    for coluna in df.columns:
        lista_valores = []
        for valor in df[coluna]:
            # Converter NaN/None para string vazia
            if pd.isna(valor):
                lista_valores.append("")
            # Converter booleanos para string
            elif isinstance(valor, bool):
                lista_valores.append(str(valor))
            # Converter números para string
            elif isinstance(valor, (int, float, np.integer, np.floating)):
                lista_valores.append(str(valor))
            # Já é string
            elif isinstance(valor, str):
                lista_valores.append(valor)
            # Qualquer outro tipo, converter para string
            else:
                lista_valores.append(str(valor) if valor is not None else "")
        
        # CORREÇÃO PRINCIPAL: str(coluna) garante que a chave seja sempre texto
        dados_serializaveis[str(coluna)] = lista_valores
    
    return dados_serializaveis

# Página inicial
@app.route('/')
def index():
    return render_template('index.html')

# Nova rota para página de configuração da planilha modelo
@app.route('/generate_excel_page')
def generate_excel_page():
    return render_template('generate_excel.html')

# Rota para gerar planilha modelo com quantidade personalizada
@app.route('/generate_excel')
def generate_excel():
    try:
        quantity = int(request.args.get('quantity', '10'))
        if quantity < 1:
            quantity = 1
        if quantity > 1000:
            quantity = 1000
    except ValueError:
        quantity = 10
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Análise Módulos FV"

    # Cabeçalhos com cores diferentes para cada grupo
    headers_row1 = [
        "Informações Gerais", "Informações Gerais", "Informações Gerais", "Informações Gerais", 
        "Informações Gerais", "Informações Gerais", "Informações Gerais", "Informações Gerais", "Informações Gerais",
        "Inspeção Visual", "Inspeção Visual", "Inspeção Visual", "Inspeção Visual", "Inspeção Visual",
        "Resistência de Isolamento", "Resistência de Isolamento", "Resistência de Isolamento", "Resistência de Isolamento", "Resistência de Isolamento",
        "Teste Curva IV", "Teste Curva IV", "Teste Curva IV", "Teste Curva IV", "Teste Curva IV", "Teste Curva IV", "Teste Curva IV", 
        "Eletroluminescência", "Eletroluminescência", "Eletroluminescência"
    ]

    # Segunda linha do cabeçalho
    headers_row2 = [
        "ID do Módulo", "NS do Módulo", "Fabricante", "Modelo", "Potência do datasheet (W)", 
        "Voc Original (V)", "Isc Original (A)", "Ano", "Bifacial/Monofacial",
        "Vidro Quebrado/Rachado?", "Backsheet Danificado?", 
        "Junction Box Danificado?", "Cabos/Conectores Danificados?",
        "Defeito Reparável?", 
        "Altura (m)", "Largura (m)",
        "Resistência Medida 1 min (MΩ)", "Resistência Medida 2 min (MΩ)", "Resistência Ôhmica Fabricante (MΩ·m²)",
        "Idade do Módulo Conhecida?",
        "Voc Medido (V)", "Isc Medido (A)", "Pmáx Medido (W)", 
        "Fill Factor Medido (%)", "Potência (% da original)",
        "Fill Factor Original (%)",
        "Foi realizado Eletroluminescência?", "Rachaduras Detectadas?", ">50% Células Danificadas?"
    ]

    # Adicionar as duas linhas de cabeçalho
    ws.append(headers_row1)
    ws.append(headers_row2)

    # Ajustar largura das colunas
    for col_idx in range(1, len(headers_row2) + 1):
        header_text = headers_row2[col_idx - 1]
        ws.column_dimensions[get_column_letter(col_idx)].width = max(20, len(str(header_text)) + 2)

    # Cores diferentes para cada grupo na primeira linha
    header_fills = {
        'Informações Gerais': PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid"),
        'Inspeção Visual': PatternFill(start_color="C0504D", end_color="C0504D", fill_type="solid"),
        'Resistência de Isolamento': PatternFill(start_color="9BBB59", end_color="9BBB59", fill_type="solid"),
        'Teste Curva IV': PatternFill(start_color="8064A2", end_color="8064A2", fill_type="solid"),
        'Eletroluminescência': PatternFill(start_color="4BACC6", end_color="4BACC6", fill_type="solid")
    }
    
    header_font = Font(bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Aplicar estilos à primeira linha
    for col_idx in range(1, len(headers_row1) + 1):
        cell = ws.cell(row=1, column=col_idx)
        header_text = headers_row1[col_idx - 1]
        
        for group, fill in header_fills.items():
            if group in header_text:
                cell.fill = fill
                break
        
        cell.font = header_font
        cell.alignment = header_alignment

    # Estilo para segunda linha
    info_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    info_font = Font(bold=True, color="000000")
    
    for col_idx in range(1, len(headers_row2) + 1):
        cell = ws.cell(row=2, column=col_idx)
        cell.fill = info_fill
        cell.font = info_font
        cell.alignment = header_alignment

    # Mesclar células das etapas
    merge_ranges = [
        ('A1:I1'), ('J1:N1'), ('O1:S1'), ('T1:AA1'), ('AB1:AD1')
    ]

    for merge_range in merge_ranges:
        ws.merge_cells(merge_range)

    ws.freeze_panes = "A3"

    # Aplicar bordas
    thin_border = Border(left=Side(style="thin"), right=Side(style="thin"),
                         top=Side(style="thin"), bottom=Side(style="thin"))
    
    for col_idx in range(1, len(headers_row1) + 1):
        ws.cell(row=1, column=col_idx).border = thin_border
    
    for col_idx in range(1, len(headers_row2) + 1):
        ws.cell(row=2, column=col_idx).border = thin_border

    # Validação de dados
    dv_sim_nao_na = DataValidation(type="list", formula1='"Sim,Não,NA"', allow_blank=True)
    for idx, header in enumerate(headers_row2, start=1):
        if "?" in header:
            col_letter = get_column_letter(idx)
            dv_sim_nao_na.add(f"{col_letter}3:{col_letter}{3+quantity}")
    ws.add_data_validation(dv_sim_nao_na)

    dv_tipo = DataValidation(type="list", formula1='"Bifacial,Monofacial"', allow_blank=True)
    col_tipo = get_column_letter(headers_row2.index("Bifacial/Monofacial") + 1)
    dv_tipo.add(f"{col_tipo}3:{col_tipo}{3+quantity}")
    ws.add_data_validation(dv_tipo)

    dv_idade = DataValidation(type="list", formula1='"Sim,Não"', allow_blank=True)
    col_idade = get_column_letter(headers_row2.index("Idade do Módulo Conhecida?") + 1)
    dv_idade.add(f"{col_idade}3:{col_idade}{3+quantity}")
    ws.add_data_validation(dv_idade)

    # Configurar proteção da planilha
    ws.protection.sheet = True
    ws.protection.enable()
    
    # ADICIONAR MÓDULOS COM FÓRMULAS - CORREÇÃO DE COLUNAS
    for i in range(quantity):
        row_data = [""] * len(headers_row2)
        row_data[0] = f"M{i+1:04d}"
        
        row_num = i + 3
        
        # COLUNAS COM FÓRMULAS - CORREÇÃO DE NÚMEROS DE COLUNA
        altura_col = 15  # O
        largura_col = 16  # P
        r1_col = 17  # Q
        r2_col = 18  # R
        resistencia_fabricante_col = 19  # S (PROTEGIDA)
        
        altura_cell = f"{get_column_letter(altura_col)}{row_num}"
        largura_cell = f"{get_column_letter(largura_col)}{row_num}"
        r1_cell = f"{get_column_letter(r1_col)}{row_num}"
        r2_cell = f"{get_column_letter(r2_col)}{row_num}"
        
        formula_resistencia = f'=IF(OR(ISNUMBER({altura_cell})=FALSE, ISNUMBER({largura_cell})=FALSE, ISNUMBER({r1_cell})=FALSE, ISNUMBER({r2_cell})=FALSE), "", ROUND(MIN({r1_cell},{r2_cell})*{altura_cell}*{largura_cell}, 2))'
        row_data[resistencia_fabricante_col - 1] = formula_resistencia
        
        pmax_col = 23  # W (Pmáx Medido - WATTS)
        pot_datasheet_col = 5  # E (Potência datasheet)
        pot_percent_col = 25  # Y (Potência % - PROTEGIDA)
        
        pmax_cell = f"{get_column_letter(pmax_col)}{row_num}"
        pot_datasheet_cell = f"{get_column_letter(pot_datasheet_col)}{row_num}"
        
        formula_potencia = f'=IF(OR(ISNUMBER({pot_datasheet_cell})=FALSE, {pot_datasheet_cell}<=0, ISNUMBER({pmax_cell})=FALSE), "", ROUND({pmax_cell}/{pot_datasheet_cell}, 4))'
        row_data[pot_percent_col - 1] = formula_potencia
        
        voc_original_col = 6  # F
        isc_original_col = 7  # G
        ff_original_col = 26  # Z (PROTEGIDA)
        
        voc_original_cell = f"{get_column_letter(voc_original_col)}{row_num}"
        isc_original_cell = f"{get_column_letter(isc_original_col)}{row_num}"
        
        formula_ff_original = f'=IF(OR(ISNUMBER({voc_original_cell})=FALSE, {voc_original_cell}<=0, ISNUMBER({isc_original_cell})=FALSE, {isc_original_cell}<=0, ISNUMBER({pot_datasheet_cell})=FALSE), "", ROUND({pot_datasheet_cell}/({voc_original_cell}*{isc_original_cell}), 4))'
        row_data[ff_original_col - 1] = formula_ff_original
        
        # DADOS DE EXEMPLO
        row_data[3] = "Exemplo"
        row_data[4] = "350"
        row_data[5] = "45.5"
        row_data[6] = "9.8"
        row_data[7] = "2015"
        row_data[8] = "Monofacial"
        row_data[14] = "1.100"
        row_data[15] = "0.650"
        row_data[16] = "150.00"
        row_data[17] = "145.00"
        row_data[19] = "Sim"
        row_data[20] = "44.2"
        row_data[21] = "9.5"
        row_data[22] = "285.00"  # W (23): Pmáx Medido em WATTS
        row_data[23] = "78.00"   # X (24): Fill Factor Medido em PORCENTAGEM
        
        ws.append(row_data)
        
        # Aplicar bordas
        for col_idx in range(1, len(headers_row2) + 1):
            ws.cell(row=row_num, column=col_idx).border = thin_border
        
        # CONFIGURAÇÃO DE FORMATO NUMÉRICO CORRIGIDA
        # Coluna W (23): Pmáx Medido (W) - NÚMERO EM WATTS
        ws.cell(row=row_num, column=23).number_format = '0.00'  # Formato numérico (ex: 285.00)
        
        # Coluna X (24): Fill Factor Medido (%) - PORCENTAGEM
        ws.cell(row=row_num, column=24).number_format = '0.00%'  # Formato porcentagem (ex: 78.00%)
        
        # Coluna Y (25): Potência (% da original) - DECIMAL
        ws.cell(row=row_num, column=25).number_format = '0.0000'  # Formato decimal (ex: 0.8143)
        
        # Coluna Z (26): Fill Factor Original (%) - PORCENTAGEM
        ws.cell(row=row_num, column=26).number_format = '0.00%'  # Formato porcentagem
        
        # Outras colunas numéricas
        ws.cell(row=row_num, column=15).number_format = '0.000'  # Altura (m) - O
        ws.cell(row=row_num, column=16).number_format = '0.000'  # Largura (m) - P
        
        # Colunas com 2 decimais (números)
        colunas_2_decimais = [5, 6, 7, 17, 18, 19, 20, 21, 22]  # E, F, G, Q, R, S, T, U, V
        for col in colunas_2_decimais:
            ws.cell(row=row_num, column=col).number_format = '0.00'
        
        # Colunas inteiras
        colunas_inteiras = [1, 8]  # A, H
        for col in colunas_inteiras:
            ws.cell(row=row_num, column=col).number_format = '0'
        
        # CONFIGURAÇÃO DE PROTEÇÃO CORRIGIDA
        todas_colunas = list(range(1, len(headers_row2) + 1))
        colunas_protegidas = [19, 25, 26]  # S, Y, Z - colunas com fórmulas
        
        for col in todas_colunas:
            cell = ws.cell(row=row_num, column=col)
            if col in colunas_protegidas:
                # PROTEGER células com fórmulas
                cell.protection = Protection(locked=True)
                cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
            else:
                # DESPROTEGER células editáveis
                cell.protection = Protection(locked=False)

    ws_instructions = wb.create_sheet(title="Instruções")
    
    instructions = [
        ["📚 INSTRUÇÕES PARA PREENCHIMENTO - CONFORME ARTIGO CIENTÍFICO"],
        ["Artigo: 'Circular solar economy: PV modules decision-making framework for reuse'"],
        ["Journal of Cleaner Production, 2023"],
        [""],
        ["ATENÇÃO IMPORTANTE PARA SUA PLANILHA:"],
        ["• O campo 'Pmáx Medido (W)' deve estar em WATTS (W), não em kW e NÃO em porcentagem"],
        ["• Exemplo: se mediu 146W, coloque 146.00 (não 0.146 e NÃO 146%)"],
        ["• O Fill Factor Medido (%) deve ser em porcentagem (ex: 78 para 78%)"],
        ["• As colunas S, Y e Z são calculadas automaticamente - NÃO EDITAR"],
        [""],
        ["FORMATAÇÃO CORRETA DAS COLUNAS:"],
        ["• Coluna W (Pmáx Medido): Número em WATTS (ex: 285.00)"],
        ["• Coluna X (Fill Factor Medido): Porcentagem (ex: 78.00%)"],
        ["• Coluna Y (Potência %): Decimal calculado (ex: 0.8143)"],
        ["• Coluna Z (Fill Factor Original): Porcentagem calculada (ex: 74.79%)"],
        [""],
        ["COLUNAS PROTEGIDAS (não editar):"],
        ["• Coluna S: Resistência Ôhmica Fabricante (MΩ·m²) - cálculo automático"],
        ["• Coluna Y: Potência (% da original) - cálculo automático"],
        ["• Coluna Z: Fill Factor Original (%) - cálculo automático"],
        [""],
        ["CRITÉRIOS TÉCNICOS DO ARTIGO:"],
        [""],
        ["1. INSPEÇÃO VISUAL (Páginas 5-6 do artigo):"],
        ["   • Vidro quebrado/rachado: RECICLAGEM IMEDIATA (♻️)"],
        ["   • Backsheet/Junction Box/Cabos danificados:"],
        ["     - Se REPARÁVEL: MANUTENÇÃO (🔧)"],
        ["     - Se NÃO REPARÁVEL: RECICLAGEM (♻️)"],
        ["   • Se NÃO HÁ DANOS: Use 'NA' em 'Defeito Reparável?'"],
        ["   • Tempo recomendado: 1 minuto por módulo"],
        [""],
        ["2. RESISTÊNCIA DE ISOLAMENTO (Página 6 do artigo):"],
        ["   • Norma: IEC 61215-2"],
        ["   • MÍNIMO ACEITÁVEL: 40 MΩ·m² (CRITÉRIO DO ARTIGO)"],
        ["   • Cálculo: Resistência Ôhmica Fabricante = MÍNIMO(R1, R2) × Altura × Largura"],
        ["   • Exemplo: Altura=1.105m, Largura=0.610m, R1=138.00, R2=124.00"],
        ["     Resultado: MÍNIMO(138,124) × 1.105 × 0.610 = 124 × 0.67405 = 83.58 MΩ·m² ✓"],
        ["   • IMPORTANTE: Use PONTO (.) como separador decimal"],
        [""],
        ["3. TESTE CURVA IV (Página 6 do artigo):"],
        ["   • Norma: IEC 62446-1"],
        ["   • Dois cenários (conforme artigo):"],
        ["     A) IDADE CONHECIDA:"],
        ["        - Degradação anual padrão: 1% (fixo no sistema)"],
        ["        - Potência Esperada = 100% - (Anos × 1%)"],
        ["        - Exemplo: 22 anos = 78% potência esperada"],
        ["        - MÍNIMO ACEITÁVEL = Esperada - 10% → 68% mínimo"],
        ["        - CLASSIFICAÇÃO:"],
        ["          * Classe A: ≥ Potência Esperada (ex: ≥78%)"],
        ["          * Classe B: ≥ Mínimo Aceitável (ex: 68-77%)"],
        ["     B) IDADE DESCONHECIDA:"],
        ["        - MÍNIMO ACEITÁVEL: 60% da potência original"],
        ["        - CLASSIFICAÇÃO:"],
        ["          * Classe A: ≥ 90%"],
        ["          * Classe B: ≥ 60% e < 90%"],
        ["   • Potência (% da original): valor decimal (ex: 0.85 para 85%)"],
        [""],
        ["4. ELETROLUMINESCÊNCIA (Página 7 do artigo):"],
        ["   • Norma: IEC TS 60904-13"],
        ["   • Critérios de rejeição:"],
        ["     - Rachaduras detectadas: RECICLAGEM"],
        ["     - >50% células danificadas: RECICLAGEM"],
        ["   • Teste opcional para validação extra"],
        [""],
        ["FÓRMULAS AUTOMÁTICAS NA PLANILHA:"],
        ["1. Resistência Ôhmica Fabricante (Coluna S):"],
        ["   =SE(OU(ÉNÚM(O3)=FALSO; ÉNÚM(P3)=FALSO; ÉNÚM(Q3)=FALSO; ÉNÚM(R3)=FALSO); \"\"; ARRED(MÍNIMO(Q3;R3)*O3*P3; 2))"],
        ["2. Potência % Original (Coluna Y):"],
        ["   =SE(OU(ÉNÚM(E3)=FALSO; E3<=0; ÉNÚM(W3)=FALSO); \"\"; ARRED(W3/E3; 4))"],
        ["3. Fill Factor Original (Coluna Z):"],
        ["   =SE(OU(ÉNÚM(F3)=FALSO; F3<=0; ÉNÚM(G3)=FALSO; G3<=0; ÉNÚM(E3)=FALSO); \"\"; ARRED(E3/(F3*G3); 4))"],
        [""],
        ["CORREÇÃO PARA SUA PLANILHA ATUAL:"],
        ["• Módulo M0001: Pmáx Medido deve ser 43.10 (WATTS), NÃO 43.10%"],
        ["• Módulo M0002: Pmáx Medido deve ser 47.00 (WATTS), NÃO 47.00%"],
        ["• Fill Factor Medido: 78.00% (não 0.78)"],
        ["• Após correção, a potência percentual será:"],
        ["   M0001: 43.10/56 = 0.7696 (76.96% - correto)"],
        ["   M0002: 47.00/56 = 0.8393 (83.93% - correto)"],
        [""],
        ["CONTATO E SUPORTE:"],
        ["• Framework desenvolvido por Marinna Pivatto"],
        ["• Sistema desenvolvido por Santiago Mateo"],
        ["• Baseado em artigo científico revisado por pares"]
    ]
    
    for instruction in instructions:
        ws_instructions.append([instruction[0]])
    
    ws_instructions.column_dimensions['A'].width = 120
    
    for row in range(1, len(instructions) + 1):
        cell = ws_instructions.cell(row=row, column=1)
        if row == 1:
            cell.font = Font(bold=True, size=14, color="FF0000")
        elif "ATENÇÃO IMPORTANTE" in str(cell.value):
            cell.font = Font(bold=True, color="FF0000", size=12)
        elif any(x in cell.value for x in ["FORMATAÇÃO CORRETA", "COLUNAS PROTEGIDAS", "CRITÉRIOS TÉCNICOS", "1. INSPEÇÃO VISUAL", "2. RESISTÊNCIA", 
                                           "3. TESTE CURVA IV", "4. ELETROLUMINESCÊNCIA"]):
            cell.font = Font(bold=True, color="0000FF")
        elif any(x in cell.value for x in ["FÓRMULAS AUTOMÁTICAS", "CORREÇÃO PARA SUA PLANILHA"]):
            cell.font = Font(bold=True, color="00AA00")
    
    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        file_path = tmp.name
        wb.save(file_path)

    return send_file(
        file_path,
        as_attachment=True,
        download_name=f"Planilha_Modulos_FV_Conforme_Artigo_{quantity}_modulos.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

# Função para verificar formato decimal
def check_decimal_format(df):
    """Verifica se há vírgulas sendo usadas como separador decimal"""
    problematic_cols = []
    for col in df.columns:
        if df[col].dtype == 'object':
            has_comma = df[col].astype(str).str.contains(r'\d,\d', na=False).any()
            if has_comma:
                problematic_cols.append(col)
    
    return problematic_cols

# Função para calcular potência esperada conforme artigo
def calcular_potencia_esperada_conforme_artigo(ano_fabricacao, idade_conhecida=True):
    """Calcula potência esperada conforme artigo (1% degradação anual fixo)"""
    try:
        ano_atual = datetime.now().year
        
        if idade_conhecida and ano_fabricacao > 0:
            idade = max(0, ano_atual - int(ano_fabricacao))
            degradacao_anual = 1.0
            potencia_esperada = 100 - (idade * degradacao_anual)
            minimo_aceitavel = max(0, potencia_esperada - 10)
            return potencia_esperada, minimo_aceitavel, idade
        else:
            return 0, 60, 0
    except:
        return 0, 0, 0

# Função principal de avaliação conforme artigo - VERSÃO CORRIGIDA
def avaliar_modulo_conforme_artigo(row, etapas_detalhadas):
    """Avalia módulo conforme critérios do artigo científico - VERSÃO CORRIGIDA"""
    
    # ==================== 1. FALHAS CRÍTICAS (RECICLAGEM IMEDIATA) ====================
    
    # 1.1. VIDRO QUEBRADO - CRÍTICO
    vidro_quebrado = str(row.get("Vidro Quebrado/Rachado?", "")).strip().upper()
    if vidro_quebrado == "SIM":
        etapas_detalhadas['visual']['fail'] += 1
        return "Reciclagem ♻️ (Vidro quebrado - Conforme artigo)"
    
    # 1.2. ELETROLUMINESCÊNCIA COM FALHAS - CRÍTICO
    el_realizado = str(row.get("Foi realizado Eletroluminescência?", "")).strip().upper()
    if el_realizado == "SIM":
        rachaduras = str(row.get("Rachaduras Detectadas?", "")).strip().upper()
        celulas_danificadas = str(row.get(">50% Células Danificadas?", "")).strip().upper()
        
        if rachaduras == "SIM" or celulas_danificadas == "SIM":
            etapas_detalhadas['el']['fail'] += 1
            return "Reciclagem ♻️ (Falha em eletroluminescência - Conforme artigo)"
        else:
            etapas_detalhadas['el']['pass'] += 1
    
    # ==================== 2. DANOS VISUAIS ====================
    
    danos_visuais = (
        str(row.get("Backsheet Danificado?", "")).strip().upper() == "SIM" or 
        str(row.get("Junction Box Danificado?", "")).strip().upper() == "SIM" or 
        str(row.get("Cabos/Conectores Danificados?", "")).strip().upper() == "SIM"
    )
    
    defeito_reparavel = str(row.get("Defeito Reparável?", "")).strip().upper()
    
    if danos_visuais:
        if defeito_reparavel == "SIM":
            etapas_detalhadas['visual']['maintenance'] += 1
            return "Manutenção 🔧 (Danos reparáveis - Conforme artigo)"
        else:
            etapas_detalhadas['visual']['fail'] += 1
            return "Reciclagem ♻️ (Danos não reparáveis - Conforme artigo)"
    else:
        etapas_detalhadas['visual']['pass'] += 1
    
    # ==================== 3. RESISTÊNCIA DE ISOLAMENTO ====================
    
    try:
        resistencia_fabricante = converter_numero(row.get("Resistência Ôhmica Fabricante (MΩ·m²)", 0))
        
        # Calcular se não estiver disponível
        if resistencia_fabricante <= 0:
            altura = converter_numero(row.get("Altura (m)", 0))
            largura = converter_numero(row.get("Largura (m)", 0))
            r1 = converter_numero(row.get("Resistência Medida 1 min (MΩ)", 0))
            r2 = converter_numero(row.get("Resistência Medida 2 min (MΩ)", 0))
            
            if altura > 0 and largura > 0 and r1 > 0 and r2 > 0:
                resistencia_fabricante = min(r1, r2) * altura * largura
        
        # CRITÉRIOS CONFORME ARTIGO CORRIGIDOS:
        # - < 40 MΩ·m²: FALHA CRÍTICA (Reciclagem)
        # - 40-60 MΩ·m²: DESEMPENHO REDUZIDO (Classe B possível)
        # - ≥ 60 MΩ·m²: EXCELENTE (Classe A possível)
        
        if resistencia_fabricante < 40:
            etapas_detalhadas['resistance']['fail'] += 1
            return f"Reciclagem ♻️ (Resistência {resistencia_fabricante:.1f} MΩ·m² < 40 MΩ·m² - Conforme artigo)"

        # Armazenar status da resistência para decisão final
        status_resistencia = "A" if resistencia_fabricante >= 60 else "B"
        etapas_detalhadas['resistance']['pass'] += 1

    except Exception as e:
        print(f"ERRO no cálculo de resistência: {e}")
        etapas_detalhadas['resistance']['fail'] += 1
        return "Reciclagem ♻️ (Erro nos dados de resistência)"
    
    # ==================== 4. TESTE CURVA IV ====================
    
    idade_conhecida = str(row.get("Idade do Módulo Conhecida?", "")).strip().upper() == "SIM"
    
    # Calcular potência percentual CORRETAMENTE
    try:
        if 'Potência (% da original)' in row and pd.notna(row.get('Potência (% da original)')):
            potencia_percent = converter_numero(row.get('Potência (% da original)', 0))
            # Converter decimal para percentual se necessário
            if potencia_percent <= 1:
                potencia_percent = potencia_percent * 100
            elif potencia_percent > 100 and potencia_percent < 1000:
                # Já está em percentual
                pass
        else:
            # Calcular manualmente
            pmax_medido = converter_numero(row.get("Pmáx Medido (W)", 0))
            potencia_datasheet = converter_numero(row.get("Potência do datasheet (W)", 0))
            
            # CORREÇÕES CRÍTICAS:
            # 1. Se Pmáx for percentual (ex: 43.10% = 0.4310)
            if 0 < pmax_medido < 1 and potencia_datasheet > 10:
                pmax_medido = pmax_medido * 100  # Converter para WATTS
            
            # 2. Se Pmáx estiver em kW (ex: 0.146 kW)
            if pmax_medido < 10 and potencia_datasheet > 10:
                pmax_medido = pmax_medido * 1000  # kW para W
            
            if potencia_datasheet > 0:
                potencia_percent = (pmax_medido / potencia_datasheet) * 100
            else:
                potencia_percent = 0
        
    except Exception as e:
        print(f"ERRO no cálculo de potência: {e}")
        potencia_percent = 0
    
    # ==================== 5. CLASSIFICAÇÃO CONFORME ARTIGO ====================
    
    if idade_conhecida:
        try:
            ano_fabricacao = converter_numero(row.get("Ano", 0))
            
            # Degradação fixa de 1% ao ano conforme artigo
            degradacao_anual = 1.0  # 1% fixo conforme artigo
            ano_atual = datetime.now().year
            idade = max(0, ano_atual - int(ano_fabricacao))
            
            # Calcular potência esperada
            potencia_esperada = max(0, 100 - (idade * degradacao_anual))
            minimo_aceitavel = max(0, potencia_esperada - 10)
            
            # VALIDAÇÃO DE DADOS
            if potencia_esperada <= 0 or minimo_aceitavel <= 0:
                # Fallback: usar critério idade desconhecida
                if potencia_percent < 60:
                    etapas_detalhadas['n_curve']['fail'] += 1
                    return "Reciclagem ♻️ (Potência <60% - Fallback)"
                else:
                    status_potencia = "A" if potencia_percent >= 90 else "B"
            else:
                # CRITÉRIOS CONFORME ARTIGO CORRIGIDOS:
                # - < mínimo aceitável: FALHA (Reciclagem)
                # - ≥ mínimo mas < esperada: DESEMPENHO REDUZIDO (Classe B)
                # - ≥ esperada: EXCELENTE (Classe A)
                
                if potencia_percent < minimo_aceitavel:
                    etapas_detalhadas['n_curve']['fail'] += 1
                    return f"Reciclagem ♻️ (Potência {potencia_percent:.1f}% < {minimo_aceitavel:.1f}% mínimo - Conforme artigo)"
                elif potencia_percent >= potencia_esperada:
                    status_potencia = "A"
                else:
                    status_potencia = "B"
                
                etapas_detalhadas['n_curve']['pass'] += 1
                
        except Exception as e:
            print(f"ERRO no cálculo idade conhecida: {e}")
            # Fallback: critério idade desconhecida
            if potencia_percent < 60:
                etapas_detalhadas['n_curve']['fail'] += 1
                return "Reciclagem ♻️ (Potência <60% - Erro no cálculo)"
            else:
                status_potencia = "A" if potencia_percent >= 90 else "B"
                etapas_detalhadas['n_curve']['pass'] += 1
                
    else:
        # IDADE DESCONHECIDA - critérios fixos do artigo
        if potencia_percent < 60:
            etapas_detalhadas['n_curve']['fail'] += 1
            return f"Reciclagem ♻️ (Potência {potencia_percent:.1f}% < 60% - Conforme artigo para idade desconhecida)"
        elif potencia_percent >= 90:
            status_potencia = "A"
        else:
            status_potencia = "B"
        
        etapas_detalhadas['n_curve']['pass'] += 1
    
    # ==================== 6. DECISÃO FINAL (CONFORME ARTIGO) ====================
    
    # CRITÉRIOS FINAIS CORRIGIDOS:
    # 1. Se ambos resistência E potência são "A" → Classe A
    # 2. Se pelo menos um é "B" (mas nenhum é "F") → Classe B
    # 3. Qualquer "F" já foi tratado acima (Reciclagem)
    
    if status_resistencia == "A" and status_potencia == "A":
        return f"Classe A ✅ (Excelente - Potência {potencia_percent:.1f}%, Resistência {resistencia_fabricante:.1f} MΩ·m²)"
    else:
        # Pelo menos um é B (mas ambos passaram nos mínimos)
        return f"Classe B ⚠️ (Aceitável - Potência {potencia_percent:.1f}%, Resistência {resistencia_fabricante:.1f} MΩ·m²)"

# Função para gerar relatório PDF completo
def generate_pdf_report(df, estatisticas, etapas_stats, resultados_lista, graph_data):
    _register_pdf_fonts()

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4,
                           rightMargin=1.5*cm, leftMargin=1.5*cm,
                           topMargin=2*cm, bottomMargin=2*cm)

    elements = []
    styles = getSampleStyleSheet()

    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Title'],
        fontSize=18,
        spaceAfter=25,
        alignment=1,
        textColor=colors.HexColor('#2C3E50'),
        fontName='Arial-Bold'
    )

    subtitle_style = ParagraphStyle(
        'Subtitle',
        parent=styles['Heading2'],
        fontSize=14,
        spaceAfter=12,
        textColor=colors.HexColor('#3498DB'),
        fontName='Arial-Bold',
        spaceBefore=20
    )

    section_style = ParagraphStyle(
        'Section',
        parent=styles['Heading3'],
        fontSize=12,
        spaceAfter=8,
        textColor=colors.HexColor('#2C3E50'),
        fontName='Arial-Bold',
        spaceBefore=15
    )

    normal_style = ParagraphStyle(
        'Normal',
        parent=styles['Normal'],
        fontSize=10,
        fontName='Arial',
        spaceAfter=6,
        leading=14
    )

    highlight_style = ParagraphStyle(
        'Highlight',
        parent=styles['Normal'],
        fontSize=10,
        fontName='Arial',
        spaceAfter=6,
        backColor=colors.HexColor('#F8F9FA'),
        borderPadding=5,
        borderColor=colors.HexColor('#DEE2E6'),
        borderWidth=1
    )

    italic_style = ParagraphStyle(
        'ItalicArial',
        parent=styles['Normal'],
        fontSize=10,
        fontName='Arial-Italic',
        spaceAfter=4
    )

    cell_style = ParagraphStyle(
        'TableCell',
        parent=styles['Normal'],
        fontName='Arial',
        fontSize=9,
        leading=12,
        wordWrap='LTR'
    )
    
    elements.append(Paragraph("<b>RELATÓRIO DE ANÁLISE TÉCNICA</b>", title_style))
    elements.append(Paragraph("<b>Módulos Fotovoltaicos - Segunda Vida</b>", subtitle_style))
    elements.append(Paragraph("<i>Conforme artigo científico: 'Circular solar economy: PV modules decision-making framework for reuse'</i>", italic_style))
    elements.append(Paragraph("<i>Journal of Cleaner Production, 2023</i>", italic_style))
    elements.append(Spacer(1, 20))
    
    data_atual = datetime.now().strftime("%d/%m/%Y %H:%M")
    info_text = f"""
    <b>Data da Análise:</b> {data_atual}<br/>
    <b>Total de Módulos Analisados:</b> {estatisticas['total_modulos']}<br/>
    <b>Sistema:</b> Estrutura de Decisão para Segunda Vida de Módulos FV<br/>
    <b>Referência:</b> Critérios técnicos baseados em artigo científico revisado por pares
    """
    elements.append(Paragraph(info_text, normal_style))
    elements.append(Spacer(1, 25))
    
    elements.append(Paragraph("<b>RESUMO EXECUTIVO</b>", subtitle_style))
    
    taxa_reuso = estatisticas['percent_classe_a'] + estatisticas['percent_classe_b']
    
    resumo_text = f"""
    A análise técnica de <b>{estatisticas['total_modulos']} módulos fotovoltaicos</b> utilizando os critérios do 
    artigo científico resultou em <b>{taxa_reuso:.1f}% de taxa de reuso total</b>. 
    <b>{estatisticas['percent_classe_a']:.1f}%</b> foram classificados como <b>Classe A</b> (potência ≥ esperada) e 
    <b>{estatisticas['percent_classe_b']:.1f}%</b> como <b>Classe B</b> (potência ≥ mínima aceitável). 
    <b>{estatisticas['percent_reciclagem']:.1f}%</b> requerem reciclagem e 
    <b>{estatisticas['percent_manutencao']:.1f}%</b> necessitam de manutenção prévia.
    """
    elements.append(Paragraph(resumo_text, normal_style))
    elements.append(Spacer(1, 15))
    
    resumo_data = [
        ["CATEGORIA", "QUANTIDADE", "PERCENTUAL", "DESCRIÇÃO"],
        ["Classe A", f"{estatisticas['classe_a']}", f"{estatisticas['percent_classe_a']}%", "Potência ≥ esperada (conforme artigo)"],
        ["Classe B", f"{estatisticas['classe_b']}", f"{estatisticas['percent_classe_b']}%", "Potência ≥ mínima aceitável"],
        ["Reciclagem", f"{estatisticas['reciclagem']}", f"{estatisticas['percent_reciclagem']}%", "Não atende critérios técnicos"],
        ["Manutenção", f"{estatisticas['manutencao']}", f"{estatisticas['percent_manutencao']}%", "Reparável para segunda vida"],
        ["TOTAL", f"{estatisticas['total_modulos']}", "100%", ""]
    ]
    
    resumo_table = Table(resumo_data, colWidths=[3.5*cm, 2.5*cm, 2.5*cm, 9.5*cm])
    resumo_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2C3E50')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Arial-Bold'),
        ('FONTNAME', (0, 1), (-1, -1), 'Arial'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('TOPPADDING', (0, 0), (-1, -1), 5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
        ('BACKGROUND', (0, 1), (-1, 1), colors.HexColor('#E8F8F5')),
        ('BACKGROUND', (0, 2), (-1, 2), colors.HexColor('#FEF9E7')),
        ('BACKGROUND', (0, 3), (-1, 3), colors.HexColor('#FDEDEC')),
        ('BACKGROUND', (0, 4), (-1, 4), colors.HexColor('#EBF5FB')),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    
    elements.append(resumo_table)
    elements.append(Spacer(1, 20))

    # Pie chart — distribuição de classificação
    elements.append(Paragraph("<b>DISTRIBUIÇÃO DE CLASSIFICAÇÃO</b>", subtitle_style))
    pie_entries = [
        (estatisticas['classe_a'],   'Classe A',   colors.HexColor('#27AE60')),
        (estatisticas['classe_b'],   'Classe B',   colors.HexColor('#F39C12')),
        (estatisticas['reciclagem'], 'Reciclagem', colors.HexColor('#E74C3C')),
        (estatisticas['manutencao'], 'Manutencao', colors.HexColor('#3498DB')),
    ]
    pie_data   = [c for c, _, _ in pie_entries if c > 0]
    pie_labels = [f"{lbl}: {c/estatisticas['total_modulos']*100:.1f}%" for c, lbl, _ in pie_entries if c > 0]
    pie_colors = [col for c, _, col in pie_entries if c > 0]
    if len(pie_data) > 1:
        drawing = Drawing(400, 190)
        pie = Pie()
        pie.x, pie.y = 80, 20
        pie.width = pie.height = 150
        pie.data   = pie_data
        pie.labels = pie_labels
        pie.slices.labelRadius = 1.25
        pie.slices.fontName    = 'Arial'
        pie.slices.fontSize    = 8
        pie.slices.strokeWidth = 0.5
        pie.slices.strokeColor = colors.white
        for i, col in enumerate(pie_colors):
            pie.slices[i].fillColor = col
        drawing.add(pie)
        elements.append(drawing)
    elements.append(Spacer(1, 15))

    elements.append(Paragraph("<b>COMPARAÇÃO COM ARTIGO CIENTÍFICO</b>", subtitle_style))
    
    comparacao_text = f"""
    <b>Artigo de referência:</b> "Circular solar economy: PV modules decision-making framework for reuse"<br/>
    <b>Publicação:</b> Journal of Cleaner Production, 2023<br/>
    <b>Método:</b> Análise de 76 módulos fotovoltaicos com 22 anos de operação<br/>
    <b>Resultado reportado:</b> <b>68% de taxa de reuso</b> para segunda vida<br/>
    <b>Nosso resultado:</b> <b>{taxa_reuso:.1f}% de taxa de reuso</b><br/>
    <b>Conclusão:</b> Resultado { 'COMPATÍVEL' if abs(taxa_reuso - 68) < 15 else 'DIVERGENTE' } com o artigo científico
    """
    elements.append(Paragraph(comparacao_text, highlight_style))
    elements.append(Spacer(1, 25))
    
    elements.append(Paragraph("<b>CRITÉRIOS TÉCNICOS APLICADOS</b>", subtitle_style))
    
    def cv(text):
        return Paragraph(text, cell_style)

    criterios_data = [
        ["ETAPA", "CRITÉRIO", "REFERÊNCIA", "VALOR"],
        ["Inspeção Visual", "Vidro quebrado", "Artigo p.5-6", cv("Reciclagem imediata")],
        ["", "Danos reparáveis", "Artigo p.5-6", cv("Manutenção")],
        ["", "Danos irreparáveis", "Artigo p.5-6", cv("Reciclagem")],
        ["Resistência Isolamento", "Mínimo aceitável", "IEC 61215-2 / Artigo p.6", cv("40 MΩ·m²")],
        ["Teste Curva IV", "Idade conhecida (1% deg/anual)", "Artigo p.6", cv("≥ (Esperada - 10%)")],
        ["", "Idade desconhecida", "Artigo p.6", cv("≥ 60% original")],
        ["", "Classificação A/B", "Artigo p.6-7", cv("Baseada em potência")],
        ["Eletroluminescência", "Rachaduras", "IEC TS 60904-13", cv("Reciclagem")],
        ["", ">50% células danificadas", "Artigo p.7", cv("Reciclagem")]
    ]
    
    criterios_table = Table(criterios_data, colWidths=[4.5*cm, 6*cm, 4.5*cm, 3*cm])
    criterios_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3498DB')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Arial-Bold'),
        ('FONTNAME', (0, 1), (-1, -1), 'Arial'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('TOPPADDING', (0, 0), (-1, -1), 5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('SPAN', (0, 1), (0, 3)),
        ('SPAN', (0, 5), (0, 7)),
        ('SPAN', (0, 8), (0, 9)),
        ('BACKGROUND', (0, 1), (-1, 3), colors.HexColor('#FEF9E7')),
        ('BACKGROUND', (0, 4), (-1, 4), colors.HexColor('#E8F6F3')),
        ('BACKGROUND', (0, 5), (-1, 7), colors.HexColor('#E8F6F3')),
        ('BACKGROUND', (0, 8), (-1, 9), colors.HexColor('#FDEDEC')),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    
    elements.append(criterios_table)
    elements.append(Spacer(1, 25))
    
    elements.append(Paragraph("<b>ANÁLISE POR ETAPA DE TESTE</b>", subtitle_style))
    
    etapas_data = [
        ["ETAPA", "APROVADOS", "REPROVADOS", "MANUTENÇÃO"],
        ["Inspeção Visual", f"{etapas_stats['visual']['pass']}%", f"{etapas_stats['visual']['fail']}%", f"{etapas_stats['visual'].get('maintenance', 0)}%"],
        ["Resistência + Curva IV", f"{etapas_stats['n_curve']['pass']}%", f"{etapas_stats['n_curve']['fail']}%", "0%"],
        ["Eletroluminescência", f"{etapas_stats['el']['pass']}%", f"{etapas_stats['el']['fail']}%", "0%"]
    ]
    
    etapas_table = Table(etapas_data, colWidths=[7*cm, 3.5*cm, 3.5*cm, 4*cm])
    etapas_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2C3E50')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Arial-Bold'),
        ('FONTNAME', (0, 1), (-1, -1), 'Arial'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('TOPPADDING', (0, 0), (-1, -1), 5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
        ('BACKGROUND', (0, 1), (-1, 1), colors.HexColor('#D5F4E6')),
        ('BACKGROUND', (0, 2), (-1, 2), colors.HexColor('#FFE5CC')),
        ('BACKGROUND', (0, 3), (-1, 3), colors.HexColor('#D6EAF8')),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    
    elements.append(etapas_table)
    elements.append(Spacer(1, 25))
    
    elements.append(Paragraph("<b>FLUXO DE DECISÃO APLICADO</b>", subtitle_style))
    
    fluxo_text = """
    1. <b>Inspeção Visual:</b> Verificação rápida (1 min/módulo) para identificar danos críticos<br/>
    2. <b>Teste de Resistência:</b> Medição conforme IEC 61215-2 (mínimo 40 MΩ·m²)<br/>
    3. <b>Teste Curva IV:</b> Avaliação de potência remanescente<br/>
       • <i>Idade conhecida:</i> Comparação com potência esperada (1% degradação anual fixo)<br/>
       • <i>Idade desconhecida:</i> Limite fixo de 60% da potência original<br/>
    4. <b>Classificação:</b> Divisão em Classe A (alta performance) e Classe B (performance aceitável)<br/>
    5. <b>Validação Opcional:</b> Eletroluminescência para detecção de defeitos internos
    """
    elements.append(Paragraph(fluxo_text, normal_style))
    elements.append(Spacer(1, 25))
    
    elements.append(Paragraph("<b>RESULTADOS COMPLETOS</b>", subtitle_style))

    if resultados_lista:
        detalhes_data = [["ID do Módulo", "Resultado da Análise"]]

        for item in resultados_lista:
            if '➝' in item:
                parts = item.split('➝', 1)
                detalhes_data.append([
                    _pdf_text(parts[0]),
                    Paragraph(_pdf_text(parts[1]), cell_style)
                ])

        detalhes_table = Table(detalhes_data, colWidths=[3.5*cm, 14.5*cm], repeatRows=1)
        detalhes_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#95A5A6')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Arial-Bold'),
            ('FONTNAME', (0, 1), (-1, -1), 'Arial'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('TOPPADDING', (0, 0), (-1, -1), 5),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
            ('GRID', (0, 0), (-1, -1), 0.25, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F2F3F4')]),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))

        elements.append(detalhes_table)
        elements.append(Spacer(1, 25))
    
    elements.append(Paragraph("<b>CONCLUSÕES E RECOMENDAÇÕES</b>", subtitle_style))
    
    conclusao_text = f"""
    1. <b>Viabilidade Técnica:</b> A taxa de reuso de <b>{taxa_reuso:.1f}%</b> demonstra o potencial de segunda vida para módulos fotovoltaicos.<br/>
    2. <b>Alinhamento Científico:</b> Os critérios aplicados estão em conformidade com literatura científica recente.<br/>
    3. <b>Sustentabilidade:</b> O reaproveitamento contribui para economia circular na indústria solar.<br/>
    4. <b>Recomendações:</b><br/>
       • Implementar sistema de rastreabilidade para módulos de segunda vida<br/>
       • Estabelecer garantias diferenciadas para Classe A e Classe B<br/>
       • Desenvolver protocolos padronizados para testes de reuso<br/>
       • Considerar aspectos econômicos na viabilidade de projetos
    """
    elements.append(Paragraph(conclusao_text, normal_style))
    elements.append(Spacer(1, 30))
    
    rodape_text = f"""
    <b>Sistema de Avaliação para Segunda Vida de Módulos Fotovoltaicos</b><br/>
    Desenvolvido com base em artigo científico revisado por pares<br/>
    Referência: Pivatto et al. (2023). Circular solar economy: PV modules decision-making framework for reuse<br/>
    Journal of Cleaner Production<br/>
    Relatório gerado em {data_atual}
    """

    elements.append(Paragraph(rodape_text, ParagraphStyle(
        'Rodape',
        parent=styles['Normal'],
        fontSize=8,
        fontName='Arial',
        textColor=colors.grey,
        alignment=1,
        spaceBefore=20
    )))
    
    doc.build(elements)
    buffer.seek(0)
    return buffer

# Função para gerar Excel com resultados
def generate_excel_with_results(df):
    wb = Workbook()
    ws = wb.active
    ws.title = "Resultados Completos"
    
    ws.append(["RESULTADOS DA ANÁLISE - CONFORME ARTIGO CIENTÍFICO"])
    ws.append(["Referência: 'Circular solar economy: PV modules decision-making framework for reuse'"])
    ws.append(["Journal of Cleaner Production, 2023"])
    ws.append([""])
    
    columns = df.columns.tolist()
    ws.append(columns)
    
    title_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    title_font = Font(bold=True, color="FFFFFF", size=14)
    
    for row in range(1, 4):
        for col in range(1, 2):
            cell = ws.cell(row=row, column=col)
            cell.fill = title_fill
            cell.font = title_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
    
    ws.merge_cells('A1:D1')
    ws.merge_cells('A2:D2')
    ws.merge_cells('A3:D3')
    
    header_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    header_font = Font(bold=True, color="000000")
    
    header_row = 5
    for col_idx, header in enumerate(columns, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = max(15, len(str(header)) + 2)
        cell = ws.cell(row=header_row, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    for idx, row in df.iterrows():
        row_data = [row[col] for col in columns]
        ws.append(row_data)
    
    if 'Resultado' in columns:
        result_col_idx = columns.index('Resultado') + 1
        
        for row in range(header_row + 1, len(df) + header_row + 1):
            cell = ws.cell(row=row, column=result_col_idx)
            result = str(cell.value) if cell.value else ""
            
            if 'Classe A' in result:
                cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                cell.font = Font(color="006100", bold=True)
            elif 'Classe B' in result:
                cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                cell.font = Font(color="9C6500", bold=True)
            elif 'Reciclagem' in result:
                cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                cell.font = Font(color="9C0006", bold=True)
            elif 'Manutenção' in result:
                cell.fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
                cell.font = Font(color="1F4E79", bold=True)
    
    ws_summary = wb.create_sheet(title="Resumo Estatístico")
    
    total = len(df)
    classe_a = len(df[df['Resultado'].str.contains('Classe A', na=False)])
    classe_b = len(df[df['Resultado'].str.contains('Classe B', na=False)])
    reciclagem = len(df[df['Resultado'].str.contains('Reciclagem', na=False)])
    manutencao = len(df[df['Resultado'].str.contains('Manutenção', na=False)])
    
    summary_data = [
        ["RESUMO DA ANÁLISE - CONFORME ARTIGO CIENTÍFICO", ""],
        ["Referência: Circular solar economy: PV modules decision-making framework for reuse", ""],
        ["Journal of Cleaner Production, 2023", ""],
        ["", ""],
        ["DADOS GERAIS", ""],
        ["Total de Módulos Analisados", total],
        ["Data da Análise", datetime.now().strftime("%d/%m/%Y %H:%M")],
        ["", ""],
        ["RESULTADOS QUANTITATIVOS", "QUANTIDADE"],
        ["Classe A (Potência ≥ Esperada)", classe_a],
        ["Classe B (Potência ≥ Mínima Aceitável)", classe_b],
        ["Reciclagem (Não atende critérios)", reciclagem],
        ["Manutenção (Reparável)", manutencao],
        ["", ""],
        ["PERCENTUAIS", ""],
        ["Taxa de Reuso Total (Classe A + B)", f"{((classe_a + classe_b) / total * 100):.1f}%"],
        ["Classe A", f"{(classe_a / total * 100):.1f}%"],
        ["Classe B", f"{(classe_b / total * 100):.1f}%"],
        ["Reciclagem", f"{(reciclagem / total * 100):.1f}%"],
        ["Manutenção", f"{(manutencao / total * 100):.1f}%"],
        ["", ""],
        ["COMPARAÇÃO COM ARTIGO", ""],
        ["Resultado do Artigo (68% reuso)", f"{'✓ COMPATÍVEL' if abs(((classe_a + classe_b) / total * 100) - 68) < 15 else '✗ DIVERGENTE'}"],
        ["", ""],
        ["CRITÉRIOS TÉCNICOS APLICADOS", ""],
        ["• Resistência Isolamento: ≥ 40 MΩ·m²", ""],
        ["• Potência (idade conhecida): ≥ (Esperada - 10%) (1% deg/anual)", ""],
        ["• Potência (idade desconhecida): ≥ 60% original", ""],
        ["• Eletroluminescência: Sem rachaduras, <50% células danificadas", ""],
    ]
    
    for row_data in summary_data:
        ws_summary.append(row_data)
    
    for col in range(1, 3):
        ws_summary.column_dimensions[get_column_letter(col)].width = 40
    
    ws_summary['A1'].font = Font(bold=True, size=14, color="FF0000")
    ws_summary['A2'].font = Font(italic=True, color="0000FF")
    
    for row in range(1, len(summary_data) + 1):
        cell = ws_summary.cell(row=row, column=1)
        if any(x in str(cell.value) for x in ["RESUMO", "DADOS GERAIS", "RESULTADOS", "PERCENTUAIS", "COMPARAÇÃO", "CRITÉRIOS"]):
            cell.font = Font(bold=True, color="0000FF")
    
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer

# Página de upload e análise
@app.route('/upload', methods=['GET', 'POST'])
def upload():
    if request.method == 'POST':
        file = request.files.get('file')
        if file and file.filename.endswith('.xlsx'):
            try:
                df = pd.read_excel(file, skiprows=1)
                
                problematic = check_decimal_format(df)
                if problematic:
                    flash(f"ATENÇÃO: As colunas {', '.join(problematic)} contêm vírgulas como separador decimal. "
                          f"Substitua por ponto (.) para evitar erros.", "warning")
                
                column_mapping = {}
                for col in df.columns:
                    col_str = str(col).strip()
                    
                    mapeamentos = {
                        'Potência do datasheet': 'Potência do datasheet (W)',
                        'Voc Original': 'Voc Original (V)',
                        'Isc Original': 'Isc Original (A)',
                        'Pmáx Medido': 'Pmáx Medido (W)',
                        'Potência (% da original)': 'Potência (% da original)',
                        'Vidro Quebrado': 'Vidro Quebrado/Rachado?',
                        'Backsheet Danificado': 'Backsheet Danificado?',
                        'Junction Box Danificado': 'Junction Box Danificado?',
                        'Cabos/Conectores Danificados': 'Cabos/Conectores Danificados?',
                        'Defeito Reparável': 'Defeito Reparável?',
                        'Bifacial/Monofacial': 'Bifacial/Monofacial',
                        'Fill Factor Medido': 'Fill Factor Medido (%)',
                        'Fill Factor Original': 'Fill Factor Original (%)',
                        'Resistência Ôhmica Fabricante': 'Resistência Ôhmica Fabricante (MΩ·m²)',
                        'Resistência Medida 1 min': 'Resistência Medida 1 min (MΩ)',
                        'Resistência Medida 2 min': 'Resistência Medida 2 min (MΩ)',
                        'Altura (m)': 'Altura (m)',
                        'Largura (m)': 'Largura (m)',
                        'Idade do Módulo Conhecida': 'Idade do Módulo Conhecida?',
                        'Foi realizado Eletroluminescência': 'Foi realizado Eletroluminescência?',
                        'Rachaduras Detectadas': 'Rachaduras Detectadas?',
                        '>50% Células Danificadas': '>50% Células Danificadas?'
                    }
                    
                    for key, value in mapeamentos.items():
                        if key in col_str:
                            column_mapping[col] = value
                            break
                
                if column_mapping:
                    df = df.rename(columns=column_mapping)
                
                colunas_necessarias = [
                    'Altura (m)', 'Largura (m)', 'Resistência Medida 1 min (MΩ)', 
                    'Resistência Medida 2 min (MΩ)', 'Resistência Ôhmica Fabricante (MΩ·m²)',
                    'Potência do datasheet (W)', 'Pmáx Medido (W)'
                ]
                
                for coluna in colunas_necessarias:
                    if coluna not in df.columns:
                        flash(f"⚠️ Coluna '{coluna}' não encontrada. Verifique o template.", "warning")
                
                etapas_detalhadas = {
                    'visual': {'pass': 0, 'fail': 0, 'maintenance': 0},
                    'resistance': {'pass': 0, 'fail': 0},
                    'n_curve': {'pass': 0, 'fail': 0},
                    'el': {'pass': 0, 'fail': 0}
                }
                
                resultados_lista = []
                dados_detalhados = []
                
                for idx, row in df.iterrows():
                    try:
                        resultado = avaliar_modulo_conforme_artigo(row, etapas_detalhadas)
                        mod_id = row.get('ID do Módulo', f'M{idx+1:03d}')
                        if pd.isna(mod_id) or mod_id == 'nan':
                            mod_id = f'M{idx+1:03d}'
                        resultados_lista.append(f"{mod_id} ➝ {resultado}")
                        
                        dados_detalhados.append({
                            'ID': str(mod_id),
                            'Resultado': resultado,
                            'Potencia_%': converter_numero(row.get('Potência (% da original)', 0)) * 100,
                            'Resistencia_MΩm²': converter_numero(row.get('Resistência Ôhmica Fabricante (MΩ·m²)', 0)),
                            'Ano': converter_numero(row.get('Ano', 0))
                        })
                        
                    except Exception as e:
                        print(f"ERRO no módulo {idx+1}: {e}")
                        resultados_lista.append(f"M{idx+1:03d} ➝ Erro na análise: {str(e)}")
                
                if resultados_lista:
                    resultados_finais = []
                    for r in resultados_lista:
                        if '➝' in r:
                            resultados_finais.append(r.split('➝')[1].strip())
                        else:
                            resultados_finais.append(r)
                    df["Resultado"] = resultados_finais
                else:
                    df["Resultado"] = "Erro na análise"
                
                total_modulos = len(df)
                classe_a = len(df[df["Resultado"].str.contains("Classe A", na=False)])
                classe_b = len(df[df["Resultado"].str.contains("Classe B", na=False)])
                reciclagem = len(df[df["Resultado"].str.contains("Reciclagem", na=False)])
                manutencao = len(df[df["Resultado"].str.contains("Manutenção", na=False)])
                
                percent_classe_a = (classe_a / total_modulos * 100) if total_modulos > 0 else 0
                percent_classe_b = (classe_b / total_modulos * 100) if total_modulos > 0 else 0
                percent_reciclagem = (reciclagem / total_modulos * 100) if total_modulos > 0 else 0
                percent_manutencao = (manutencao / total_modulos * 100) if total_modulos > 0 else 0
                
                estatisticas = {
                    'total_modulos': total_modulos,
                    'classe_a': classe_a,
                    'classe_b': classe_b,
                    'reciclagem': reciclagem,
                    'manutencao': manutencao,
                    'percent_classe_a': round(percent_classe_a, 1),
                    'percent_classe_b': round(percent_classe_b, 1),
                    'percent_reciclagem': round(percent_reciclagem, 1),
                    'percent_manutencao': round(percent_manutencao, 1)
                }
                
                etapas_stats = {}
                for etapa, valores in etapas_detalhadas.items():
                    total_etapa = sum(valores.values())
                    if total_etapa > 0:
                        etapas_stats[etapa] = {
                            'pass': round((valores['pass'] / total_etapa) * 100, 1),
                            'fail': round((valores['fail'] / total_etapa) * 100, 1),
                            'maintenance': round((valores.get('maintenance', 0) / total_etapa) * 100, 1)
                        }
                    else:
                        etapas_stats[etapa] = {'pass': 0, 'fail': 0, 'maintenance': 0}
                
                total_pass = classe_a + classe_b

                vi_pass         = etapas_detalhadas['visual']['pass'] + etapas_detalhadas['visual'].get('maintenance', 0)
                resistance_pass = etapas_detalhadas['resistance']['pass']
                iv_pass         = etapas_detalhadas['n_curve']['pass']   # passes both resistance + IV curve = total_pass

                t = total_modulos if total_modulos > 0 else 1
                vi_pass_pct         = vi_pass         / t * 100
                resistance_pass_pct = resistance_pass / t * 100
                iv_pass_pct         = total_pass      / t * 100   # use total_pass (classe A+B) as ground truth

                vi_fail_pct         = 100              - vi_pass_pct
                resistance_fail_pct = vi_pass_pct      - resistance_pass_pct
                iv_fail_pct         = resistance_pass_pct - iv_pass_pct

                graph_data = {
                    'labels': [
                        'Total Descomissionado',
                        'Inspeção Visual',
                        'Teste de Resistência',
                        'Curva IV / Potência'
                    ],
                    'pass_values': [
                        round(100,                 1),
                        round(vi_pass_pct,         1),
                        round(resistance_pass_pct, 1),
                        round(iv_pass_pct,         1),
                    ],
                    # fail = cumulative discard so that green + red = 100% on every bar
                    'fail_values': [
                        0,
                        round(-(100 - vi_pass_pct),         1),
                        round(-(100 - resistance_pass_pct), 1),
                        round(-(100 - iv_pass_pct),         1),
                    ],
                    'quantidades': {
                        'total':            total_modulos,
                        'vi_pass':          vi_pass,
                        'vi_fail':          total_modulos - vi_pass,
                        'resistance_pass':  resistance_pass,
                        'resistance_fail':  vi_pass - resistance_pass,
                        'iv_pass':          total_pass,
                        'iv_fail':          resistance_pass - total_pass,
                        'total_pass':       total_pass,
                    }
                }
                
                # CORREÇÃO AQUI: Chamada da função segura
                df_data_seguro = dataframe_para_dict_serializavel_seguro(df)
                
                session['df_data'] = df_data_seguro
                session['estatisticas'] = estatisticas
                session['etapas_stats'] = etapas_stats
                session['resultados_lista'] = resultados_lista
                session['graph_data'] = graph_data
                session['dados_detalhados'] = dados_detalhados
                
                taxa_reuso = estatisticas['percent_classe_a'] + estatisticas['percent_classe_b']
                
                problemas = []
                if 'Pmáx Medido (W)' in df.columns:
                    for idx, valor in enumerate(df['Pmáx Medido (W)']):
                        try:
                            val_num = converter_numero(valor)
                            # Verificar se está como porcentagem (ex: 43.10% = 0.4310)
                            if 0 < val_num < 1:
                                problemas.append(f"Módulo {idx+1}: Pmáx Medido = {valor} (parece porcentagem, deve ser em WATTS)")
                            elif val_num < 10 and val_num > 0:
                                problemas.append(f"Módulo {idx+1}: Pmáx Medido = {valor}W (pode estar em kW)")
                        except:
                            pass
                
                if problemas:
                    flash("⚠️ ATENÇÃO: Possível erro nas medições:", "warning")
                    for problema in problemas:
                        flash(f"   • {problema}", "warning")
                    flash("   Corrija para Watts (ex: 43.10% → 43.10 WATTS)", "warning")
                
                if taxa_reuso > 0:
                    flash(f"✅ Análise concluída! Taxa de reuso: {taxa_reuso:.1f}%", "success")
                    flash(f"📊 Resultado do artigo: 68% reuso para módulos de 22 anos", "info")
                    flash(f"🔧 {estatisticas['manutencao']} módulos necessitam de manutenção prévia", "warning")
                    flash(f"📈 {estatisticas['classe_a']} módulos Classe A, {estatisticas['classe_b']} módulos Classe B", "info")
                
                return render_template("resultado.html", 
                                     resultados=resultados_lista,
                                     total_modulos=total_modulos,
                                     estatisticas=estatisticas,
                                     etapas_stats=etapas_stats,
                                     graph_data=graph_data)
            
            except Exception as e:
                error_msg = f"❌ Erro ao processar o arquivo: {str(e)}"
                flash(error_msg, "error")
                app.logger.error(f"Erro no upload: {traceback.format_exc()}")
                print(f"ERRO CRÍTICO: {traceback.format_exc()}")
                return render_template("upload.html")
        
        else:
            flash("⚠️ Por favor, envie um arquivo Excel (.xlsx) válido.", "warning")
    
    return render_template("upload.html")

# Rota para baixar Excel com resultados
@app.route('/download_excel', methods=['POST'])
def download_excel():
    try:
        if 'df_data' not in session:
            return "Dados não encontrados. Por favor, realize uma análise primeiro.", 400
        
        df_data = session['df_data']
        df = pd.DataFrame(df_data)
        
        if 'Resultado' not in df.columns:
            return "Coluna de Resultado não encontrada.", 400
        
        excel_buffer = generate_excel_with_results(df)
        
        filename = f"Resultados_Analise_FV_Artigo_{datetime.now().strftime('%Y-%m-%d_%H.%M.%S')}.xlsx"
        
        return send_file(
            excel_buffer,
            as_attachment=True,
            download_name=filename,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    except Exception as e:
        app.logger.error(f"Erro no download Excel: {traceback.format_exc()}")
        return f"Erro ao gerar Excel: {str(e)}", 500

# Rota para baixar PDF com resultados
@app.route('/download_pdf', methods=['POST'])
def download_pdf():
    try:
        if 'df_data' not in session or 'estatisticas' not in session:
            return "Dados não encontrados. Por favor, realize uma análise primeiro.", 400
        
        df_data = session['df_data']
        df = pd.DataFrame(df_data)
        estatisticas = session['estatisticas']
        etapas_stats = session['etapas_stats']
        resultados_lista = session.get('resultados_lista', [])
        graph_data = session.get('graph_data', {})
        
        pdf_buffer = generate_pdf_report(df, estatisticas, etapas_stats, resultados_lista, graph_data)
        
        filename = f"Relatorio_Analise_FV_Artigo_{datetime.now().strftime('%Y-%m-%d_%H.%M.%S')}.pdf"
        
        return send_file(
            pdf_buffer,
            as_attachment=True,
            download_name=filename,
            mimetype="application/pdf"
        )
    except Exception as e:
        app.logger.error(f"Erro no download PDF: {traceback.format_exc()}")
        return f"Erro ao gerar PDF: {str(e)}", 500

# Rota para API de análise
@app.route('/api/analyze', methods=['POST'])
def api_analyze():
    try:
        if 'file' not in request.files:
            return jsonify({'error': 'Nenhum arquivo enviado'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'error': 'Nome de arquivo vazio'}), 400
        
        if not file.filename.endswith('.xlsx'):
            return jsonify({'error': 'Formato inválido. Use .xlsx'}), 400
        
        df = pd.read_excel(file, skiprows=1)
        
        etapas_detalhadas = {
            'visual': {'pass': 0, 'fail': 0, 'maintenance': 0},
            'resistance': {'pass': 0, 'fail': 0},
            'n_curve': {'pass': 0, 'fail': 0},
            'el': {'pass': 0, 'fail': 0}
        }
        
        resultados = []
        for idx, row in df.iterrows():
            resultado = avaliar_modulo_conforme_artigo(row, etapas_detalhadas)
            resultados.append({
                'id': row.get('ID do Módulo', f'M{idx+1:03d}'),
                'resultado': resultado
            })
        
        total = len(resultados)
        classe_a = sum(1 for r in resultados if 'Classe A' in r['resultado'])
        classe_b = sum(1 for r in resultados if 'Classe B' in r['resultado'])
        reciclagem = sum(1 for r in resultados if 'Reciclagem' in r['resultado'])
        manutencao = sum(1 for r in resultados if 'Manutenção' in r['resultado'])
        
        return jsonify({
            'success': True,
            'total_modulos': total,
            'classe_a': classe_a,
            'classe_b': classe_b,
            'reciclagem': reciclagem,
            'manutencao': manutencao,
            'taxa_reuso': ((classe_a + classe_b) / total * 100) if total > 0 else 0,
            'resultados': resultados,
            'etapas': etapas_detalhadas
        })
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# Página de documentação do sistema
@app.route('/documentacao')
def documentacao():
    return render_template('documentacao.html')

# Rota para limpar sessão
@app.route('/limpar')
def limpar():
    session.clear()
    flash("Sessão limpa com sucesso!", "info")
    return redirect('/')

# Manipuladores de erro
@app.errorhandler(404)
def page_not_found(e):
    return render_template('404.html'), 404

@app.errorhandler(500)
def internal_server_error(e):
    return render_template('500.html'), 500

if __name__ == '__main__':
    app.run(debug=True, port=5000, host='0.0.0.0')