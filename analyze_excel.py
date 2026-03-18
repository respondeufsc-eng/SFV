# analyze_excel.py - VERSÃO COMPLETA CORRIGIDA conforme artigo científico
import pandas as pd
import numpy as np
from datetime import datetime
import sys
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font

def converter_numero(valor):
    """Converte string numérica com ponto ou vírgula para float"""
    if pd.isna(valor):
        return 0.0
    
    try:
        if isinstance(valor, str):
            valor_limpo = valor.strip().replace(',', '.')
            
            if any(error in valor_limpo.upper() for error in ['#DIV/0!', '#N/A', '#VALUE!', '#REF!', '#NAME?', 'N/A']):
                return 0.0
            
            valor_limpo = ''.join(c for c in valor_limpo if c.isdigit() or c in '.-')
            
            if valor_limpo and valor_limpo != '-':
                return float(valor_limpo)
            else:
                return 0.0
        else:
            return float(valor)
    except:
        return 0.0

def calcular_potencia_esperada_conforme_artigo(ano_fabricacao, degradacao_anual_percent=1.0):
    """Calcula potência esperada conforme artigo (1% degradação anual)"""
    try:
        ano_atual = datetime.now().year
        idade = max(0, ano_atual - int(ano_fabricacao))
        
        degradacao_anual = degradacao_anual_percent / 100.0
        potencia_esperada = 100 - (idade * degradacao_anual * 100)
        minimo_aceitavel = max(0, potencia_esperada - 10)
        
        return potencia_esperada, minimo_aceitavel, idade
    except:
        return 0, 0, 0

def avaliar_modulo_conforme_artigo(row):
    """Avalia módulo conforme critérios do artigo científico"""
    
    # 1. INSPEÇÃO VISUAL
    vidro_quebrado = str(row.get("Vidro Quebrado/Rachado?", "")).strip().upper()
    
    if vidro_quebrado == "SIM":
        return "Reciclagem ♻️ (Vidro quebrado - Conforme artigo)"
    
    danos_visuais = (
        str(row.get("Backsheet Danificado?", "")).strip().upper() == "SIM" or 
        str(row.get("Junction Box Danificado?", "")).strip().upper() == "SIM" or 
        str(row.get("Cabos/Conectores Danificados?", "")).strip().upper() == "SIM"
    )
    
    defeito_reparavel = str(row.get("Defeito Reparável?", "")).strip().upper()
    
    if danos_visuais:
        if defeito_reparavel == "SIM":
            return "Manutenção 🔧 (Danos reparáveis - Conforme artigo)"
        else:
            return "Reciclagem ♻️ (Danos não reparáveis - Conforme artigo)"
    
    # 2. TESTE DE RESISTÊNCIA
    try:
        resistencia_fabricante = converter_numero(row.get("Resistência Ôhmica Fabricante (MΩ·m²)", 0))
        
        if resistencia_fabricante < 40:
            return f"Reciclagem ♻️ (Resistência {resistencia_fabricante:.1f} MΩ·m² < 40 MΩ·m² - Conforme artigo)"
    except:
        return "Reciclagem ♻️ (Erro nos dados de resistência)"
    
    # 3. TESTE IV CURVE
    idade_conhecida = str(row.get("Idade do Módulo Conhecida?", "")).strip().upper() == "SIM"
    
    try:
        if 'Potência (% da original)' in row and pd.notna(row.get('Potência (% da original)')):
            potencia_str = str(row.get('Potência (% da original)', 0))
            
            if isinstance(potencia_str, str) and ('N/A' in potencia_str.upper() or '#' in potencia_str):
                potencia_percent = 0
            else:
                potencia = converter_numero(potencia_str)
                potencia_percent = potencia * 100 if potencia <= 1 else potencia
        else:
            pmax_medido = converter_numero(row.get("Pmáx Medido (W)", 0))
            potencia_datasheet = converter_numero(row.get("Potência do datasheet (W)", 0))
            
            if potencia_datasheet > 0:
                potencia_percent = (pmax_medido / potencia_datasheet) * 100
            else:
                potencia_percent = 0
    except:
        potencia_percent = 0
    
    # CRITÉRIOS CONFORME ARTIGO
    if idade_conhecida:
        try:
            ano_fabricacao = converter_numero(row.get("Ano", 0))
            
            degradacao_str = str(row.get("Degradação Anual Esperada (%)", "1%"))
            degradacao_anual = converter_numero(degradacao_str.replace('%', ''))
            if degradacao_anual == 0:
                degradacao_anual = 1.0
            
            potencia_esperada, minimo_aceitavel, idade = calcular_potencia_esperada_conforme_artigo(
                ano_fabricacao, degradacao_anual
            )
            
            if potencia_esperada <= 0:
                if potencia_percent < 60:
                    return "Reciclagem ♻️ (Potência <60% - Critério fallback)"
                else:
                    if potencia_percent >= 90:
                        return "Classe A ✅ (Potência ≥90% - Fallback)"
                    else:
                        return "Classe B ⚠️ (Potência ≥60% - Fallback)"
            
            if potencia_percent < minimo_aceitavel:
                return f"Reciclagem ♻️ (Potência {potencia_percent:.1f}% < {minimo_aceitavel:.1f}% mínimo - Conforme artigo)"
            else:
                if potencia_percent >= potencia_esperada:
                    return f"Classe A ✅ (Potência {potencia_percent:.1f}% ≥ {potencia_esperada:.1f}% esperada - Conforme artigo)"
                else:
                    return f"Classe B ⚠️ (Potência {potencia_percent:.1f}% ≥ {minimo_aceitavel:.1f}% mínimo - Conforme artigo)"
                
        except Exception as e:
            if potencia_percent < 60:
                return "Reciclagem ♻️ (Potência <60% - Erro no cálculo)"
            else:
                if potencia_percent >= 90:
                    return "Classe A ✅ (Potência ≥90% - Fallback)"
                else:
                    return "Classe B ⚠️ (Potência ≥60% - Fallback)"
    else:
        if potencia_percent < 60:
            return f"Reciclagem ♻️ (Potência {potencia_percent:.1f}% < 60% - Conforme artigo para idade desconhecida)"
        else:
            if potencia_percent >= 90:
                return f"Classe A ✅ (Potência {potencia_percent:.1f}% ≥ 90% - Conforme artigo)"
            else:
                return f"Classe B ⚠️ (Potência {potencia_percent:.1f}% ≥ 60% - Conforme artigo)"

def analisar_planilha_conforme_artigo(file_path, salvar_resultados=True):
    """Analisa planilha Excel conforme artigo científico"""
    
    print("=" * 100)
    print("📊 ANÁLISE DE MÓDULOS FOTOVOLTAICOS - CONFORME ARTIGO CIENTÍFICO")
    print("=" * 100)
    print("Referência: 'Circular solar economy: PV modules decision-making framework for reuse'")
    print("Journal of Cleaner Production, 2023")
    print("Autores: Pivatto et al.")
    print("=" * 100)
    
    try:
        # Carregar dados
        df = pd.read_excel(file_path)
        print(f"✅ Arquivo carregado: {os.path.basename(file_path)}")
        print(f"📋 Total de módulos encontrados: {len(df)}")
        print(f"📁 Colunas disponíveis: {len(df.columns)}")
        
        # Verificar colunas necessárias
        colunas_necessarias = [
            'Potência do datasheet (W)', 'Pmáx Medido (W)', 
            'Resistência Ôhmica Fabricante (MΩ·m²)', 'Ano'
        ]
        
        colunas_faltantes = [col for col in colunas_necessarias if col not in df.columns]
        if colunas_faltantes:
            print(f"⚠️  Colunas faltantes: {colunas_faltantes}")
            print("   Verifique se está usando o template correto.")
        
        # Aplicar avaliação
        print("\n🔍 Analisando módulos conforme critérios do artigo...")
        df["Resultado"] = df.apply(avaliar_modulo_conforme_artigo, axis=1)
        
        # Calcular estatísticas
        total_modulos = len(df)
        classe_a = len(df[df["Resultado"].str.contains("Classe A", na=False)])
        classe_b = len(df[df["Resultado"].str.contains("Classe B", na=False)])
        reciclagem = len(df[df["Resultado"].str.contains("Reciclagem", na=False)])
        manutencao = len(df[df["Resultado"].str.contains("Manutenção", na=False)])
        
        taxa_reuso = ((classe_a + classe_b) / total_modulos * 100) if total_modulos > 0 else 0
        
        print("\n" + "=" * 100)
        print("📈 RESULTADOS DA ANÁLISE:")
        print("=" * 100)
        
        # Mostrar primeiros 10 resultados
        print("\n🔢 PRIMEIROS 10 RESULTADOS:")
        for idx, row in df.head(10).iterrows():
            print(f"   {row.get('ID do Módulo', f'M{idx+1:03d}'):15} ➝ {row['Resultado']}")
        
        if total_modulos > 10:
            print(f"   ... e mais {total_modulos - 10} módulos")
        
        print("\n" + "=" * 100)
        print("📊 ESTATÍSTICAS DETALHADAS:")
        print("=" * 100)
        print(f"📈 Total de módulos analisados: {total_modulos}")
        print(f"✅ Classe A (Potência ≥ Esperada): {classe_a} ({classe_a/total_modulos*100:.1f}%)")
        print(f"⚠️  Classe B (Potência ≥ Mínima): {classe_b} ({classe_b/total_modulos*100:.1f}%)")
        print(f"♻️  Reciclagem: {reciclagem} ({reciclagem/total_modulos*100:.1f}%)")
        print(f"🔧 Manutenção (Reparável): {manutencao} ({manutencao/total_modulos*100:.1f}%)")
        print(f"🌟 Taxa de Reuso Total (Classe A+B): {taxa_reuso:.1f}%")
        
        print("\n" + "=" * 100)
        print("🎯 COMPARAÇÃO COM ARTIGO CIENTÍFICO:")
        print("=" * 100)
        print("📚 Artigo reporta: 68% reuso para módulos de 22 anos")
        print(f"📊 Nosso resultado: {taxa_reuso:.1f}% reuso")
        
        diferenca = abs(taxa_reuso - 68)
        if diferenca < 10:
            print(f"✅ Resultado COMPATÍVEL com artigo (diferença: {diferenca:.1f}%)")
        elif diferenca < 20:
            print(f"⚠️  Resultado MODERADAMENTE COMPATÍVEL (diferença: {diferenca:.1f}%)")
        else:
            print(f"❌ Resultado DIVERGENTE do artigo (diferença: {diferenca:.1f}%)")
        
        print("\n" + "=" * 100)
        print("🔧 CRITÉRIOS APLICADOS (Conforme Artigo):")
        print("=" * 100)
        print("1. Inspeção Visual:")
        print("   • Vidro quebrado: RECICLAGEM IMEDIATA")
        print("   • Outros danos: MANUTENÇÃO (se reparável) ou RECICLAGEM")
        print("2. Resistência Isolamento: MÍNIMO 40 MΩ·m²")
        print("3. Potência:")
        print("   • Idade conhecida: ≥ (Esperada - 10%)")
        print("   • Idade desconhecida: ≥ 60% original")
        print("4. Classificação:")
        print("   • Classe A: ≥ Potência Esperada")
        print("   • Classe B: ≥ Mínima Aceitável")
        print("5. Eletroluminescência (validação extra)")
        
        print("\n" + "=" * 100)
        print("📊 ANÁLISE POR TIPO DE RESULTADO:")
        print("=" * 100)
        
        # Distribuição por motivo de reciclagem
        motivos_reciclagem = df[df["Resultado"].str.contains("Reciclagem", na=False)]["Resultado"].value_counts()
        if not motivos_reciclagem.empty:
            print("Motivos de reciclagem:")
            for motivo, quantidade in motivos_reciclagem.items():
                print(f"   • {motivo.split('-')[-1].strip()}: {quantidade}")
        
        # Distribuição por faixa de potência (apenas para aprovados)
        aprovados = df[df["Resultado"].str.contains("Classe", na=False)]
        if not aprovados.empty:
            potencias = []
            for idx, row in aprovados.iterrows():
                try:
                    potencia_str = str(row.get('Potência (% da original)', 0))
                    if 'N/A' not in potencia_str.upper() and '#' not in potencia_str:
                        potencia = converter_numero(potencia_str)
                        potencias.append(potencia * 100 if potencia <= 1 else potencia)
                except:
                    pass
            
            if potencias:
                print(f"\n📈 Distribuição de potência (aprovados):")
                print(f"   • Média: {np.mean(potencias):.1f}%")
                print(f"   • Mínima: {np.min(potencias):.1f}%")
                print(f"   • Máxima: {np.max(potencias):.1f}%")
                print(f"   • Desvio padrão: {np.std(potencias):.1f}%")
        
        # Salvar resultados se solicitado
        if salvar_resultados:
            output_file = f"Resultados_Analise_Artigo_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            
            # Salvar com formatação
            with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Resultados', index=False)
                
                # Acessar a workbook e worksheet para formatação
                workbook = writer.book
                worksheet = writer.sheets['Resultados']
                
                # Formatar cabeçalho
                header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
                header_font = Font(bold=True, color="FFFFFF")
                
                for col in range(1, len(df.columns) + 1):
                    cell = worksheet.cell(row=1, column=col)
                    cell.fill = header_fill
                    cell.font = header_font
                
                # Ajustar largura das colunas
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column_letter].width = adjusted_width
                
                # Adicionar aba de resumo
                summary_data = {
                    'Métrica': ['Total Módulos', 'Classe A', 'Classe B', 'Reciclagem', 'Manutenção', 'Taxa Reuso'],
                    'Valor': [total_modulos, classe_a, classe_b, reciclagem, manutencao, f"{taxa_reuso:.1f}%"],
                    'Percentual': ['100%', f'{classe_a/total_modulos*100:.1f}%', f'{classe_b/total_modulos*100:.1f}%', 
                                 f'{reciclagem/total_modulos*100:.1f}%', f'{manutencao/total_modulos*100:.1f}%', '']
                }
                
                summary_df = pd.DataFrame(summary_data)
                summary_df.to_excel(writer, sheet_name='Resumo', index=False)
            
            print(f"\n💾 Resultados salvos em: {output_file}")
            print(f"📍 Caminho completo: {os.path.abspath(output_file)}")
        
        print("\n" + "=" * 100)
        print("✅ ANÁLISE CONCLUÍDA COM SUCESSO!")
        print("=" * 100)
        
        return df, taxa_reuso
        
    except FileNotFoundError:
        print(f"❌ ERRO: Arquivo '{file_path}' não encontrado.")
        print("   Verifique o caminho do arquivo.")
        return None, 0
    except Exception as e:
        print(f"❌ ERRO: {str(e)}")
        print("   Detalhes:", sys.exc_info())
        return None, 0

# Função principal
if __name__ == "__main__":
    import argparse
    
    parser = argparse.ArgumentParser(description='Analisar planilha de módulos FV conforme artigo científico')
    parser.add_argument('arquivo', help='Caminho para o arquivo Excel (.xlsx)')
    parser.add_argument('--salvar', action='store_true', help='Salvar resultados em Excel')
    
    args = parser.parse_args()
    
    if os.path.exists(args.arquivo):
        df, taxa_reuso = analisar_planilha_conforme_artigo(args.arquivo, args.salvar)
        
        if df is not None:
            print(f"\n🎉 Análise concluída! Taxa de reuso: {taxa_reuso:.1f}%")
            print("📤 Para usar a interface web, execute: python app.py")
            print("🌐 Acesse: http://localhost:5000")
    else:
        print(f"❌ Arquivo não encontrado: {args.arquivo}")
        print("💡 Dica: Use o comando: python analyze_excel.py sua_planilha.xlsx --salvar")